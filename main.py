import itertools
import click
from openpyxl import Workbook, load_workbook


@click.command()
@click.option(
    "--file", required=True, help="Location of the file that should be chunked."
)
@click.option(
    "--sheet-name", required=True, help="Name of the worksheet that should be read."
)
@click.option(
    "--chunk-size", default=1000, help="The max size of the newly created batch files."
)
def chunk(file: str, sheet_name: str, chunk_size: int):
    workbook = load_workbook(file)

    worksheet = workbook[sheet_name]
    row_group = batched(worksheet.iter_rows(), chunk_size)
    for index, rows in enumerate(row_group):
        new_workbook = Workbook()
        new_worksheet = new_workbook.create_sheet(sheet_name)
        for row in rows:
            values = [cell.value for cell in row]
            new_worksheet.append(values)
        new_workbook.save(f"chunks/chunk-{index + 1}.xlsx")


def batched(iterable, batch_size):
    if batch_size < 1:
        raise ValueError("Batch size must be a positive integer.")
    it = iter(iterable)
    while batch := tuple(itertools.islice(it, batch_size)):
        yield batch


if __name__ == "__main__":
    chunk()
